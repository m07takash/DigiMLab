"""
DigiMLab — Streamlit Web UI
テストセット選択 → 実行 → 結果閲覧・分析
"""
import asyncio
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import yaml

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
CONFIG_DIR = Path(__file__).parent / "config"
DATASETS_DIR = Path(__file__).parent / "datasets"
REPORTS_DIR = Path(__file__).parent / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

OCEAN_LABELS = {
    "O": "開放性 (Openness)",
    "C": "誠実性 (Conscientiousness)",
    "E": "外向性 (Extraversion)",
    "A": "協調性 (Agreeableness)",
    "N": "神経症傾向 (Neuroticism)",
}

# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def load_yaml(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_models_config() -> dict:
    return load_yaml(CONFIG_DIR / "models.yaml")


def load_settings() -> dict:
    return load_yaml(CONFIG_DIR / "settings.yaml")


def load_dataset_questions(path: str) -> list[dict]:
    full = Path(__file__).parent / path
    if not full.exists():
        return []
    with open(full, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# API情報取得（エージェント一覧・エンジン一覧等）
# ---------------------------------------------------------------------------

def _format_duration(started_at: str, finished_at: str | None = None) -> str:
    """開始〜終了（または現在）の経過時間を 'Xm Ys' 形式で返す"""
    try:
        start = datetime.fromisoformat(started_at)
        end = datetime.fromisoformat(finished_at) if finished_at else datetime.now()
        delta = end - start
        total_sec = int(delta.total_seconds())
        if total_sec < 60:
            return f"{total_sec}秒"
        minutes, seconds = divmod(total_sec, 60)
        if minutes < 60:
            return f"{minutes}分{seconds}秒"
        hours, minutes = divmod(minutes, 60)
        return f"{hours}時間{minutes}分{seconds}秒"
    except Exception:
        return ""


def _api_base_url(model_config: dict) -> str:
    """base_urlからAPIのベースパスを取得（末尾の /run 等を除去）"""
    url = model_config.get("base_url", "")
    if url.endswith("/run"):
        return url[:-4]
    return url.rstrip("/")


@st.cache_data(ttl=300)
def fetch_health(base_url: str) -> dict:
    import requests
    try:
        resp = requests.get(f"{base_url}/health", timeout=30)
        return resp.json()
    except Exception as e:
        return {"status": "error", "detail": str(e)}


@st.cache_data(ttl=300)
def fetch_agents(base_url: str) -> list[dict]:
    import requests
    try:
        resp = requests.get(f"{base_url}/agents", timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            return data.get("agents", data) if isinstance(data, dict) else data
    except Exception:
        pass
    return []


@st.cache_data(ttl=300)
def fetch_engines(base_url: str, agent_file: str) -> tuple[list[str], str]:
    """エンジン一覧とデフォルトエンジンを返す"""
    import requests
    try:
        resp = requests.get(f"{base_url}/agents/{agent_file}/engines", timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            # {"LLM": {"default": "...", "engines": [...]}} 形式
            if isinstance(data, dict) and "LLM" in data:
                llm = data["LLM"]
                engines = llm.get("engines", [])
                default = llm.get("default", "")
                return engines, default
            # フォールバック: {"engines": [...]} 形式
            if isinstance(data, dict) and "engines" in data:
                return data["engines"], ""
            if isinstance(data, list):
                return data, ""
    except Exception:
        pass
    return [], ""


@st.cache_data(ttl=300)
def fetch_web_search_engines(base_url: str) -> list[str]:
    import requests
    try:
        resp = requests.get(f"{base_url}/web_search_engines", timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            if isinstance(data, dict) and "engines" in data:
                return data["engines"]
            if isinstance(data, list):
                return data
    except Exception:
        pass
    return []


def save_results_json(results: list[dict], run_id: str) -> Path:
    out = REPORTS_DIR / f"run_{run_id}.json"
    out.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def generate_pdf(results: list[dict], summary: dict, meta: dict, job_type: str, run_id: str) -> bytes:
    """結果をPDFとして生成し、バイト列を返す"""
    import tempfile
    from fpdf import FPDF

    FONT_PATH = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
    FONT_BOLD_PATH = "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc"

    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    # フォント登録
    pdf.add_font("JP", "", FONT_PATH)
    pdf.add_font("JP", "B", FONT_BOLD_PATH)

    # --- 表紙 ---
    pdf.add_page()
    pdf.set_font("JP", "B", 24)
    pdf.ln(30)
    pdf.cell(0, 15, "DigiMLab - テスト結果レポート", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    pdf.set_font("JP", "", 12)
    for k, v in meta.items():
        pdf.cell(0, 8, f"{k}: {v}", new_x="LMARGIN", new_y="NEXT")

    valid_count = sum(1 for r in results if not r.get("is_error"))
    error_count = sum(1 for r in results if r.get("is_error"))
    pdf.cell(0, 8, f"有効回答数: {valid_count}問", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"エラー数: {error_count}問", new_x="LMARGIN", new_y="NEXT")

    # --- OCEAN分析（MPIの場合） ---
    if job_type == "mpi" and summary:
        pdf.add_page()
        pdf.set_font("JP", "B", 16)
        pdf.cell(0, 12, "OCEAN分析", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # レーダーチャートを画像として埋め込み
        try:
            radar_fig = render_radar_chart(summary)
            bar_fig = render_bar_chart(summary)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f_radar:
                radar_fig.write_image(f_radar.name, width=500, height=400, scale=2)
                radar_path = f_radar.name
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f_bar:
                bar_fig.write_image(f_bar.name, width=500, height=400, scale=2)
                bar_path = f_bar.name
            pdf.image(radar_path, x=10, w=130)
            pdf.image(bar_path, x=150, w=130)
            import os as _os
            _os.unlink(radar_path)
            _os.unlink(bar_path)
        except Exception:
            pdf.set_font("JP", "", 10)
            pdf.cell(0, 8, "(グラフ画像の生成に失敗しました)", new_x="LMARGIN", new_y="NEXT")

        # 因子別サマリテーブル
        pdf.ln(10)
        pdf.set_font("JP", "B", 14)
        pdf.cell(0, 10, "因子別サマリ", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        pdf.set_font("JP", "B", 10)
        col_widths = [20, 60, 25, 25, 25, 35]
        headers = ["因子", "因子名", "総問数", "有効回答", "エラー", "平均スコア"]
        for w, h in zip(col_widths, headers):
            pdf.cell(w, 8, h, border=1, align="C")
        pdf.ln()

        pdf.set_font("JP", "", 10)
        for f_key, label in OCEAN_LABELS.items():
            d = summary.get(f_key, {})
            mean_str = f"{d['mean']:.4f}" if d.get("mean") is not None else "N/A"
            row = [f_key, label, str(d.get("total", 0)), str(d.get("valid", 0)),
                   str(d.get("errors", 0)), mean_str]
            for w, val in zip(col_widths, row):
                pdf.cell(w, 7, val, border=1, align="C")
            pdf.ln()

    # --- 回答明細 ---
    pdf.add_page()
    pdf.set_font("JP", "B", 14)
    pdf.cell(0, 10, "回答明細", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    if job_type == "mpi":
        col_widths = [15, 70, 12, 12, 60, 15, 15, 18, 15]
        headers = ["ID", "質問文", "因子", "キー", "AIの回答", "評点", "調整後", "正規化", "エラー"]
    else:
        col_widths = [15, 80, 30, 120, 20]
        headers = ["No.", "質問", "カテゴリ", "AIの回答", "エラー"]

    pdf.set_font("JP", "B", 8)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 7, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("JP", "", 7)
    for r in results:
        if job_type == "mpi":
            row = [
                str(r.get("id", "")),
                str(r.get("text", ""))[:35],
                str(r.get("factor", "")),
                str(r.get("keyed", "")),
                str(r.get("raw_response", ""))[:30],
                str(r.get("rating", "-")),
                str(r.get("adjusted", "-")),
                f"{r['normalized']:.2f}" if r.get("normalized") is not None else "-",
                "Yes" if r.get("is_error") else "",
            ]
        else:
            row = [
                str(r.get("id", "")),
                str(r.get("question", ""))[:40],
                str(r.get("category", ""))[:15],
                str(r.get("raw_response", ""))[:60],
                "Yes" if r.get("is_error") else "",
            ]
        for w, val in zip(col_widths, row):
            pdf.cell(w, 6, val, border=1)
        pdf.ln()

    return bytes(pdf.output())


def save_results_excel(results: list[dict], summary: dict, meta: dict, run_id: str) -> Path:
    """結果をExcelに保存"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    err_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center = Alignment(horizontal="center")

    # Sheet 1: 回答一覧
    ws = wb.active
    ws.title = "回答一覧"
    headers = ["No.", "ID", "質問文", "OCEAN因子", "因子名", "キー方向",
               "AIの回答(raw)", "評点(1-5)", "調整後", "正規化(0-1)", "状態"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = thin

    for i, r in enumerate(results, 2):
        is_err = r.get("is_error", False)
        vals = [
            i - 1, r.get("id", ""), r.get("text", ""), r.get("factor", ""),
            OCEAN_LABELS.get(r.get("factor", ""), ""), r.get("keyed", ""),
            r.get("raw_response", ""),
            r.get("rating") if r.get("rating") is not None else "-",
            r.get("adjusted") if r.get("adjusted") is not None else "-",
            round(r["normalized"], 2) if r.get("normalized") is not None else "-",
            "エラー" if is_err else "OK",
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = thin
            if col >= 8:
                c.alignment = center
            if is_err:
                c.fill = err_fill

    for col_letter, w in [("A", 5), ("B", 8), ("C", 35), ("D", 10), ("E", 28),
                           ("F", 10), ("G", 18), ("H", 10), ("I", 8), ("J", 12), ("K", 8)]:
        ws.column_dimensions[col_letter].width = w

    # Sheet 2: OCEANサマリ
    ws2 = wb.create_sheet("OCEANサマリ")
    headers2 = ["因子", "因子名", "総問数", "有効回答数", "エラー数", "平均スコア(正規化)"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = thin

    for i, (f, label) in enumerate(OCEAN_LABELS.items(), 2):
        d = summary.get(f, {})
        vals = [f, label, d.get("total", 0), d.get("valid", 0),
                d.get("errors", 0), d.get("mean", "N/A")]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=v)
            c.border = thin
            if col >= 3:
                c.alignment = center

    for col_letter, w in [("A", 8), ("B", 30), ("C", 10), ("D", 12), ("E", 10), ("F", 18)]:
        ws2.column_dimensions[col_letter].width = w

    # Sheet 3: テスト情報
    ws3 = wb.create_sheet("テスト情報")
    for i, (k, v) in enumerate(meta.items(), 1):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=i, column=1).border = thin
        ws3.cell(row=i, column=2, value=str(v)).border = thin
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 65

    out = REPORTS_DIR / f"run_{run_id}.xlsx"
    wb.save(out)
    return out


def compute_ocean_summary(results: list[dict]) -> dict:
    """結果リストからOCEAN因子別のサマリを計算"""
    factor_data = defaultdict(lambda: {"total": 0, "valid": 0, "errors": 0, "scores": []})
    for r in results:
        f = r.get("factor", "")
        if f not in OCEAN_LABELS:
            continue
        factor_data[f]["total"] += 1
        if r.get("normalized") is not None:
            factor_data[f]["valid"] += 1
            factor_data[f]["scores"].append(r["normalized"])
        else:
            factor_data[f]["errors"] += 1

    summary = {}
    for f in OCEAN_LABELS:
        d = factor_data[f]
        mean = round(sum(d["scores"]) / len(d["scores"]), 4) if d["scores"] else None
        summary[f] = {
            "total": d["total"],
            "valid": d["valid"],
            "errors": d["errors"],
            "mean": mean,
        }
    return summary


def render_radar_chart(summary: dict) -> go.Figure:
    """OCEANレーダーチャート"""
    factors = list(OCEAN_LABELS.keys())
    labels = [OCEAN_LABELS[f] for f in factors]
    values = [summary[f]["mean"] if summary[f]["mean"] is not None else 0 for f in factors]
    # レーダーチャートは閉じる
    values_closed = values + [values[0]]
    labels_closed = labels + [labels[0]]

    fig = go.Figure(data=go.Scatterpolar(
        r=values_closed,
        theta=labels_closed,
        fill="toself",
        name="OCEAN",
        line=dict(color="#4472C4"),
    ))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 1])),
        showlegend=False,
        margin=dict(l=80, r=80, t=40, b=40),
        height=400,
    )
    return fig


def render_bar_chart(summary: dict) -> go.Figure:
    """OCEAN棒グラフ"""
    factors = list(OCEAN_LABELS.keys())
    labels = [OCEAN_LABELS[f] for f in factors]
    values = [summary[f]["mean"] if summary[f]["mean"] is not None else 0 for f in factors]
    colors = ["#4472C4", "#ED7D31", "#70AD47", "#FFC000", "#5B9BD5"]

    fig = go.Figure(data=go.Bar(
        x=labels, y=values,
        marker_color=colors,
        text=[f"{v:.2f}" if v else "N/A" for v in values],
        textposition="outside",
    ))
    fig.update_layout(
        yaxis=dict(range=[0, 1.1], title="正規化スコア"),
        margin=dict(l=40, r=40, t=40, b=80),
        height=400,
    )
    return fig


# ---------------------------------------------------------------------------
# 結果表示（共通）
# ---------------------------------------------------------------------------

def _render_result_viewer(key_prefix: str):
    """完了済みジョブを選択して結果を表示する共通ウィジェット"""
    completed = [j for j in list_jobs() if j["state"] in ("completed", "stopped")
                 and load_job_type(j["job_id"]) not in ("rp_judge", "style_judge")]
    if not completed:
        st.info("完了済みのテスト結果はありません。")
        return

    job_labels = {}
    for j in completed:
        m = j.get("meta", {})
        job_labels[j["job_id"]] = (
            f"{m.get('テスト名', j['job_id'])} | "
            f"{m.get('対象モデル', '')} | "
            f"{m.get('agent_file', '')} | "
            f"{j.get('started_at', '')[:19]}"
        )

    sel_id = st.selectbox(
        "結果を選択", list(job_labels.keys()),
        format_func=lambda k: job_labels[k],
        key=f"result_sel_{key_prefix}",
    )

    results = load_job_results(sel_id)
    if not results:
        return

    jtype = load_job_type(sel_id)
    meta = load_job_meta(sel_id)
    valid_count = sum(1 for r in results if not r.get("is_error"))
    error_count = sum(1 for r in results if r.get("is_error"))

    # テスト情報
    job_status = load_job_status(sel_id)
    duration = ""
    if job_status:
        duration = _format_duration(
            job_status.get("started_at", ""),
            job_status.get("finished_at") or None,
        )

    with st.expander("テスト情報", expanded=False):
        for k, v in meta.items():
            st.text(f"{k}: {v}")
        st.text(f"有効回答数: {valid_count}問")
        st.text(f"エラー数: {error_count}問")
        if duration:
            st.text(f"実行時間: {duration}")

    if jtype == "mpi":
        # --- MPI: OCEAN分析 ---
        ocean_summary = compute_ocean_summary(results)

        st.subheader("OCEAN分析")
        col_r, col_b = st.columns(2)
        with col_r:
            st.plotly_chart(render_radar_chart(ocean_summary), use_container_width=True, key=f"radar_{key_prefix}_{sel_id}")
        with col_b:
            st.plotly_chart(render_bar_chart(ocean_summary), use_container_width=True, key=f"bar_{key_prefix}_{sel_id}")

        st.subheader("因子別サマリ")
        rows = []
        for fk, lbl in OCEAN_LABELS.items():
            d = ocean_summary.get(fk, {})
            rows.append({
                "因子": fk, "因子名": lbl,
                "総問数": d.get("total", 0), "有効回答": d.get("valid", 0),
                "エラー": d.get("errors", 0),
                "平均スコア": f"{d['mean']:.4f}" if d.get("mean") is not None else "N/A",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        st.subheader("回答明細")
        df = pd.DataFrame(results)
        cols = ["id", "text", "factor", "keyed", "raw_response", "rating", "adjusted", "normalized", "is_error"]
        cols = [c for c in cols if c in df.columns]
        df_display = df[cols].rename(columns={
            "id": "ID", "text": "質問文", "factor": "因子", "keyed": "キー",
            "raw_response": "AIの回答", "rating": "評点", "adjusted": "調整後",
            "normalized": "正規化", "is_error": "エラー",
        })
    elif jtype == "psycho":
        # --- AIPsychoBench: 多尺度心理測定 ---
        df_ps = pd.DataFrame(results)
        valid_ps = df_ps[df_ps["is_error"] == False]

        # 尺度別サマリ
        st.subheader("尺度別分析")
        if not valid_ps.empty:
            scale_summary = valid_ps.groupby("scale").agg(
                total=("normalized", "count"),
                mean=("normalized", "mean"),
            ).reset_index()
            scale_summary["mean"] = scale_summary["mean"].round(4)
            scale_summary = scale_summary.rename(columns={"scale": "尺度", "total": "有効回答", "mean": "平均スコア(正規化)"})

            # 棒グラフ
            fig_ps = go.Figure(data=go.Bar(
                x=scale_summary["尺度"],
                y=scale_summary["平均スコア(正規化)"],
                text=[f"{v:.2f}" for v in scale_summary["平均スコア(正規化)"]],
                textposition="outside",
            ))
            fig_ps.update_layout(
                yaxis=dict(range=[0, 1.1], title="正規化スコア (0-1)"),
                margin=dict(l=40, r=40, t=40, b=100), height=450,
                xaxis_tickangle=-45,
            )
            st.plotly_chart(fig_ps, use_container_width=True, key=f"psycho_bar_{key_prefix}_{sel_id}")
            st.dataframe(scale_summary, use_container_width=True, hide_index=True)

            # カテゴリ別（展開可能）
            with st.expander("カテゴリ別詳細", expanded=False):
                cat_summary = valid_ps.groupby(["scale", "category"]).agg(
                    total=("normalized", "count"),
                    mean=("normalized", "mean"),
                ).reset_index()
                cat_summary["mean"] = cat_summary["mean"].round(4)
                cat_summary = cat_summary.rename(columns={
                    "scale": "尺度", "category": "カテゴリ",
                    "total": "有効回答", "mean": "平均スコア",
                })
                st.dataframe(cat_summary, use_container_width=True, hide_index=True)

        st.subheader("回答明細")
        cols_ps = ["id", "scale", "category", "text", "raw_response", "rating", "adjusted", "normalized", "is_error"]
        cols_ps = [c for c in cols_ps if c in df_ps.columns]
        df_display = df_ps[cols_ps].rename(columns={
            "id": "ID", "scale": "尺度", "category": "カテゴリ", "text": "質問文",
            "raw_response": "AIの回答", "rating": "評点", "adjusted": "調整後",
            "normalized": "正規化", "is_error": "エラー",
        })
    elif jtype == "rp":
        # --- RP Bench: ロールプレイ一覧 ---
        has_judge = any(r.get("judge_scores") for r in results)

        st.subheader("サマリ")
        col_m1, col_m2, col_m3 = st.columns(3)
        col_m1.metric("総シナリオ数", len(results))
        col_m2.metric("回答取得", valid_count)
        col_m3.metric("エラー", error_count)

        # --- Judge採点結果（ある場合） ---
        if has_judge:
            scored = [r for r in results if r.get("judge_scores")]
            axes = ["consistency", "naturalness", "relevance", "persona_accuracy"]
            axes_labels = {
                "consistency": "一貫性",
                "naturalness": "自然さ",
                "relevance": "適切さ",
                "persona_accuracy": "ペルソナ精度",
            }
            # 軸別平均
            axis_avgs = {}
            for ax in axes:
                vals = [r["judge_scores"][ax] for r in scored if ax in r.get("judge_scores", {})]
                axis_avgs[ax] = round(sum(vals) / len(vals), 2) if vals else 0

            st.subheader("Judge採点分析")
            # レーダーチャート
            col_jr, col_jb = st.columns(2)
            with col_jr:
                labels = [axes_labels[a] for a in axes]
                values = [axis_avgs[a] for a in axes]
                values_closed = values + [values[0]]
                labels_closed = labels + [labels[0]]
                fig_r = go.Figure(data=go.Scatterpolar(
                    r=values_closed, theta=labels_closed, fill="toself",
                    line=dict(color="#4472C4"),
                ))
                fig_r.update_layout(
                    polar=dict(radialaxis=dict(visible=True, range=[0, 5])),
                    showlegend=False, margin=dict(l=80, r=80, t=40, b=40), height=400,
                )
                st.plotly_chart(fig_r, use_container_width=True, key=f"rp_radar_{key_prefix}_{sel_id}")
            with col_jb:
                fig_b = go.Figure(data=go.Bar(
                    x=[axes_labels[a] for a in axes],
                    y=[axis_avgs[a] for a in axes],
                    marker_color=["#4472C4", "#ED7D31", "#70AD47", "#FFC000"],
                    text=[f"{axis_avgs[a]:.2f}" for a in axes],
                    textposition="outside",
                ))
                fig_b.update_layout(
                    yaxis=dict(range=[0, 5.5], title="スコア (1-5)"),
                    margin=dict(l=40, r=40, t=40, b=80), height=400,
                )
                st.plotly_chart(fig_b, use_container_width=True, key=f"rp_bar_{key_prefix}_{sel_id}")

            # 軸別サマリテーブル
            summary_rows = []
            for ax in axes:
                summary_rows.append({
                    "観点": axes_labels[ax],
                    "平均スコア": f"{axis_avgs[ax]:.2f}",
                    "採点数": len([r for r in scored if ax in r.get("judge_scores", {})]),
                })
            overall_avg = round(sum(axis_avgs.values()) / len(axis_avgs), 2)
            summary_rows.append({"観点": "総合平均", "平均スコア": f"{overall_avg:.2f}", "採点数": len(scored)})
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

        # --- Judge採点実行UI ---
        def _rp_judge_ui(label: str, description: str):
            """RP Judge採点のUI部品（初回/再実行共通）"""
            st.caption(description)
            try:
                from env_loader import env
                _api_key = env("OPENAI_API_KEY", default="")
                _default_model = env("JUDGE_MODEL", default="gpt-5-mini")
            except Exception:
                _api_key = ""
                _default_model = "gpt-5-mini"

            _model = st.text_input("Judgeモデル", value=_default_model, key=f"judge_model_{key_prefix}_{sel_id}_{label}")
            if st.button(f"Judge採点を{label}", key=f"run_judge_{key_prefix}_{sel_id}_{label}", type="primary"):
                if not _api_key:
                    st.error("system.env に OPENAI_API_KEY が設定されていません。")
                else:
                    import subprocess
                    _jid = f"judge_{sel_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    _jdir = JOBS_DIR / _jid
                    _jdir.mkdir(parents=True, exist_ok=True)
                    _jdata = {
                        "job_id": _jid, "job_type": "rp_judge", "source_job_id": sel_id,
                        "judge_config": {"api_key": _api_key, "model": _model},
                        "delay_seconds": 1.0, "max_retries": 3,
                        "meta": {"テスト名": f"Judge採点 ({sel_id})", "Judgeモデル": _model},
                        "_started_at": datetime.now().isoformat(),
                    }
                    (_jdir / "job.json").write_text(json.dumps(_jdata, ensure_ascii=False, indent=2), encoding="utf-8")
                    (_jdir / "status.json").write_text(
                        json.dumps({"state": "starting", "total": 0, "completed": 0}, ensure_ascii=False), encoding="utf-8"
                    )
                    _runner = Path(__file__).parent / "job_runner.py"
                    _proc = subprocess.Popen(
                        [sys.executable, str(_runner), _jid], cwd=str(Path(__file__).parent),
                        stdout=open(str(_jdir / "stdout.log"), "w"),
                        stderr=open(str(_jdir / "stderr.log"), "w"),
                        start_new_session=True,
                    )
                    (_jdir / "pid").write_text(str(_proc.pid))
                    st.success(f"Judge採点ジョブを開始しました: {_jid}")
                    st.rerun()

        st.divider()
        if not has_judge:
            st.subheader("LLM-as-Judge 採点")
            _rp_judge_ui("実行", "回答収集済みの結果に対して、LLMが4軸（一貫性・自然さ・適切さ・ペルソナ精度）で採点します。")
        else:
            with st.expander("Judge採点を再実行", expanded=False):
                _rp_judge_ui("再実行", "既存の採点結果を上書きして再採点します。")

        st.divider()
        st.subheader("ロールプレイ結果")
        # 各シナリオを展開表示
        for r in results:
            persona_label = r.get("persona", r.get("id", ""))
            icon = "✅" if not r.get("is_error") else "❌"
            judge_badge = ""
            if r.get("judge_scores"):
                avg = r.get("judge_avg", 0)
                judge_badge = f" | Judge: {avg*5:.1f}/5"
            with st.expander(f"{icon} {r.get('id','')} — {persona_label}{judge_badge}", expanded=False):
                st.markdown(f"**ペルソナ:** {r.get('persona', '')}")
                st.caption(r.get("persona_description", ""))
                st.markdown(f"**ユーザー発話:** {r.get('user_input', '')}")
                st.markdown(f"**AIの応答:**")
                st.info(r.get("raw_response", ""))
                if r.get("reference"):
                    st.markdown(f"**参考回答:**")
                    st.caption(r.get("reference", ""))
                if r.get("judge_scores"):
                    st.markdown("**Judge採点:**")
                    js = r["judge_scores"]
                    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                    col_s1.metric("一貫性", f"{js.get('consistency', '-')}/5")
                    col_s2.metric("自然さ", f"{js.get('naturalness', '-')}/5")
                    col_s3.metric("適切さ", f"{js.get('relevance', '-')}/5")
                    col_s4.metric("ペルソナ精度", f"{js.get('persona_accuracy', '-')}/5")
                    if r.get("judge_reason"):
                        st.caption(f"**根拠:** {r['judge_reason']}")

        st.subheader("回答一覧（テーブル）")
        df = pd.DataFrame(results)
        cols = ["id", "persona", "user_input", "raw_response", "reference", "is_error"]
        if has_judge:
            cols += ["judge_avg", "judge_reason"]
        cols = [c for c in cols if c in df.columns]
        rename_map = {
            "id": "ID", "persona": "ペルソナ", "user_input": "ユーザー発話",
            "raw_response": "AIの応答", "reference": "参考回答", "is_error": "エラー",
            "judge_avg": "Judge平均", "judge_reason": "採点根拠",
        }
        df_display = df[cols].rename(columns=rename_map)
    elif jtype == "style":
        # --- Style: 執筆スタイル再現度 ---
        has_judge = any(r.get("judge_scores") for r in results)

        st.subheader("サマリ")
        col_m1, col_m2, col_m3 = st.columns(3)
        col_m1.metric("総サンプル数", len(results))
        col_m2.metric("回答取得", valid_count)
        col_m3.metric("エラー", error_count)

        # --- Judge採点結果（ある場合） ---
        if has_judge:
            scored = [r for r in results if r.get("judge_scores")]
            axes = ["substance", "vocabulary", "tone", "coherence"]
            axes_labels = {
                "substance": "内容の再現性",
                "vocabulary": "語彙の再現性",
                "tone": "口調・文体の再現性",
                "coherence": "論理構成の再現性",
            }
            axis_avgs = {}
            for ax in axes:
                vals = [r["judge_scores"][ax] for r in scored if ax in r.get("judge_scores", {})]
                axis_avgs[ax] = round(sum(vals) / len(vals), 2) if vals else 0

            st.subheader("Judge採点分析")
            col_sr, col_sb = st.columns(2)
            with col_sr:
                labels = [axes_labels[a] for a in axes]
                values = [axis_avgs[a] for a in axes]
                values_closed = values + [values[0]]
                labels_closed = labels + [labels[0]]
                fig_r = go.Figure(data=go.Scatterpolar(
                    r=values_closed, theta=labels_closed, fill="toself",
                    line=dict(color="#70AD47"),
                ))
                fig_r.update_layout(
                    polar=dict(radialaxis=dict(visible=True, range=[0, 5])),
                    showlegend=False, margin=dict(l=80, r=80, t=40, b=40), height=400,
                )
                st.plotly_chart(fig_r, use_container_width=True, key=f"style_radar_{key_prefix}_{sel_id}")
            with col_sb:
                fig_b = go.Figure(data=go.Bar(
                    x=[axes_labels[a] for a in axes],
                    y=[axis_avgs[a] for a in axes],
                    marker_color=["#70AD47", "#4472C4", "#ED7D31", "#FFC000"],
                    text=[f"{axis_avgs[a]:.2f}" for a in axes],
                    textposition="outside",
                ))
                fig_b.update_layout(
                    yaxis=dict(range=[0, 5.5], title="スコア (1-5)"),
                    margin=dict(l=40, r=40, t=40, b=80), height=400,
                )
                st.plotly_chart(fig_b, use_container_width=True, key=f"style_bar_{key_prefix}_{sel_id}")

            summary_rows = []
            for ax in axes:
                summary_rows.append({
                    "観点": axes_labels[ax],
                    "平均スコア": f"{axis_avgs[ax]:.2f}",
                    "採点数": len([r for r in scored if ax in r.get("judge_scores", {})]),
                })
            overall_avg = round(sum(axis_avgs.values()) / len(axis_avgs), 2)
            summary_rows.append({"観点": "総合平均", "平均スコア": f"{overall_avg:.2f}", "採点数": len(scored)})
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

        # --- Style Judge採点実行UI ---
        def _style_judge_ui(label: str, description: str):
            st.caption(description)
            try:
                from env_loader import env
                _api_key = env("OPENAI_API_KEY", default="")
                _default_model = env("JUDGE_MODEL", default="gpt-5-mini")
            except Exception:
                _api_key = ""
                _default_model = "gpt-5-mini"

            _model = st.text_input("Judgeモデル", value=_default_model, key=f"style_judge_model_{key_prefix}_{sel_id}_{label}")
            if st.button(f"Judge採点を{label}", key=f"run_style_judge_{key_prefix}_{sel_id}_{label}", type="primary"):
                if not _api_key:
                    st.error("system.env に OPENAI_API_KEY が設定されていません。")
                else:
                    import subprocess
                    _jid = f"style_judge_{sel_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    _jdir = JOBS_DIR / _jid
                    _jdir.mkdir(parents=True, exist_ok=True)
                    _jdata = {
                        "job_id": _jid, "job_type": "style_judge", "source_job_id": sel_id,
                        "judge_config": {"api_key": _api_key, "model": _model},
                        "delay_seconds": 1.0, "max_retries": 3,
                        "meta": {"テスト名": f"Style Judge ({sel_id})", "Judgeモデル": _model},
                        "_started_at": datetime.now().isoformat(),
                    }
                    (_jdir / "job.json").write_text(json.dumps(_jdata, ensure_ascii=False, indent=2), encoding="utf-8")
                    (_jdir / "status.json").write_text(
                        json.dumps({"state": "starting", "total": 0, "completed": 0}, ensure_ascii=False), encoding="utf-8"
                    )
                    _runner = Path(__file__).parent / "job_runner.py"
                    _proc = subprocess.Popen(
                        [sys.executable, str(_runner), _jid], cwd=str(Path(__file__).parent),
                        stdout=open(str(_jdir / "stdout.log"), "w"),
                        stderr=open(str(_jdir / "stderr.log"), "w"),
                        start_new_session=True,
                    )
                    (_jdir / "pid").write_text(str(_proc.pid))
                    st.success(f"Style Judge採点ジョブを開始しました: {_jid}")
                    st.rerun()

        st.divider()
        if not has_judge:
            st.subheader("LLM-as-Judge 採点")
            _style_judge_ui("実行", "回答収集済みの結果に対して、LLMが4軸（内容・語彙・口調・論理構成の再現性）で採点します。")
        else:
            with st.expander("Judge採点を再実行", expanded=False):
                _style_judge_ui("再実行", "既存の採点結果を上書きして再採点します。")

        st.divider()
        st.subheader("スタイル再現結果")
        for r in results:
            person_label = r.get("person", r.get("id", ""))
            icon = "✅" if not r.get("is_error") else "❌"
            judge_badge = ""
            if r.get("judge_scores"):
                avg = r.get("judge_avg", 0)
                judge_badge = f" | Judge: {avg*5:.1f}/5"
            with st.expander(f"{icon} {r.get('id','')} — {person_label}{judge_badge}", expanded=False):
                st.markdown(f"**著者:** {r.get('person', '')}")
                st.markdown(f"**お題:** {r.get('prompt', '')}")
                if r.get("writing_samples"):
                    with st.expander("過去の文章サンプル", expanded=False):
                        for i_s, s in enumerate(r["writing_samples"]):
                            st.caption(f"サンプル{i_s+1}: {s}")
                st.markdown(f"**AIの文章:**")
                st.info(r.get("raw_response", ""))
                if r.get("reference"):
                    st.markdown(f"**参考文章:**")
                    st.caption(r.get("reference", ""))
                if r.get("judge_scores"):
                    st.markdown("**Judge採点:**")
                    js = r["judge_scores"]
                    col_js1, col_js2, col_js3, col_js4 = st.columns(4)
                    col_js1.metric("内容", f"{js.get('substance', '-')}/5")
                    col_js2.metric("語彙", f"{js.get('vocabulary', '-')}/5")
                    col_js3.metric("口調", f"{js.get('tone', '-')}/5")
                    col_js4.metric("論理構成", f"{js.get('coherence', '-')}/5")

        st.subheader("回答一覧（テーブル）")
        df = pd.DataFrame(results)
        cols = ["id", "person", "prompt", "raw_response", "is_error"]
        if has_judge:
            cols += ["judge_avg"]
        cols = [c for c in cols if c in df.columns]
        df_display = df[cols].rename(columns={
            "id": "ID", "person": "著者", "prompt": "お題",
            "raw_response": "AIの文章", "is_error": "エラー", "judge_avg": "Judge平均",
        })
    else:
        # --- カスタムQ&A: 質問と回答の一覧 ---
        st.subheader("サマリ")
        st.metric("総質問数", len(results))
        col_m1, col_m2 = st.columns(2)
        col_m1.metric("回答取得", valid_count)
        col_m2.metric("エラー", error_count)

        st.subheader("質問・回答一覧")
        df = pd.DataFrame(results)
        cols = ["id", "question", "category", "raw_response", "is_error"]
        cols = [c for c in cols if c in df.columns]
        df_display = df[cols].rename(columns={
            "id": "No.", "question": "質問", "category": "カテゴリ",
            "raw_response": "AIの回答", "is_error": "エラー",
        })

    st.dataframe(df_display, use_container_width=True, hide_index=True, height=500)

    # ダウンロード
    st.divider()
    col_d1, col_d2, col_d3, col_d4 = st.columns(4)
    with col_d1:
        if jtype == "mpi":
            pdf_summary = compute_ocean_summary(results)
        else:
            pdf_summary = {}
        pdf_data = generate_pdf(results, pdf_summary, meta, jtype, sel_id)
        st.download_button(
            "PDF ダウンロード",
            data=pdf_data,
            file_name=f"report_{sel_id}.pdf",
            mime="application/pdf",
            key=f"pdf_{key_prefix}_{sel_id}",
        )
    with col_d2:
        st.download_button(
            "CSV ダウンロード",
            data=df_display.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"result_{sel_id}.csv",
            mime="text/csv",
            key=f"csv_{key_prefix}_{sel_id}",
        )
    with col_d3:
        if jtype == "mpi":
            xl_path = save_results_excel(results, compute_ocean_summary(results), meta, sel_id)
        else:
            xl_path = save_results_excel(results, {}, meta, sel_id)
        with open(xl_path, "rb") as fxl:
            st.download_button(
                "Excel ダウンロード",
                data=fxl.read(),
                file_name=xl_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"xlsx_{key_prefix}_{sel_id}",
            )
    with col_d4:
        st.download_button(
            "JSON ダウンロード",
            data=json.dumps(results, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name=f"result_{sel_id}.json",
            mime="application/json",
            key=f"json_{key_prefix}_{sel_id}",
        )


# ---------------------------------------------------------------------------
# API呼び出し（テスト実行）
# ---------------------------------------------------------------------------

async def call_custom_api(
    base_url: str,
    user_input: str,
    session_id: str,
    params: dict,
) -> dict:
    """カスタムAPI(DigitalMATSUMOTO等)を呼び出す"""
    import aiohttp

    body = {
        "service_info": params.get("service_info", {"SERVICE_ID": "DigiMLab", "SERVICE_DATA": {}}),
        "user_info": params.get("user_info", {"USER_ID": "anonymous", "USER_DATA": {}}),
        "session_id": session_id,
        "user_input": user_input,
    }
    # 基本パラメータ
    for key in ("agent_file", "engine", "situation"):
        if key in params:
            body[key] = params[key]
    # boolパラメータ
    for key in ("stream_mode", "save_digest", "memory_use",
                "magic_word_use", "meta_search", "rag_query_gene"):
        if key in params and params[key] is not None:
            body[key] = params[key]
    # web_search関連
    if params.get("web_search"):
        body["web_search"] = True
        if "web_search_engine" in params:
            body["web_search_engine"] = params["web_search_engine"]
    elif "web_search" in params:
        body["web_search"] = False

    timeout = aiohttp.ClientTimeout(total=120)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.post(base_url, json=body, headers={"Content-Type": "application/json"}) as resp:
            if resp.status != 200:
                return {"error": f"HTTP {resp.status}", "response": ""}
            return await resp.json()


# ---------------------------------------------------------------------------
# ジョブ管理（バックグラウンド実行）
# ---------------------------------------------------------------------------

JOBS_DIR = REPORTS_DIR / "jobs"
JOBS_DIR.mkdir(exist_ok=True)


def create_job(
    job_id: str,
    questions: list[dict],
    model_config: dict,
    system_prompt: str,
    session_id: str,
    delay_seconds: float,
    max_retries: int,
    meta: dict,
    job_type: str = "mpi",
) -> Path:
    """ジョブ定義ファイルを作成し、バックグラウンドプロセスを起動"""
    import subprocess

    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    job_data = {
        "job_id": job_id,
        "job_type": job_type,
        "questions": questions,
        "model_config": model_config,
        "system_prompt": system_prompt,
        "session_id": session_id,
        "delay_seconds": delay_seconds,
        "max_retries": max_retries,
        "meta": meta,
        "_started_at": datetime.now().isoformat(),
    }
    (job_dir / "job.json").write_text(
        json.dumps(job_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # 初期ステータス
    (job_dir / "status.json").write_text(
        json.dumps({"state": "starting", "total": len(questions), "completed": 0}, ensure_ascii=False),
        encoding="utf-8",
    )

    # バックグラウンドでjob_runner.pyを起動
    runner_path = Path(__file__).parent / "job_runner.py"
    proc = subprocess.Popen(
        [sys.executable, str(runner_path), job_id],
        cwd=str(Path(__file__).parent),
        stdout=open(str(job_dir / "stdout.log"), "w"),
        stderr=open(str(job_dir / "stderr.log"), "w"),
        start_new_session=True,
    )
    # PIDを保存（停止用）
    (job_dir / "pid").write_text(str(proc.pid))
    return job_dir


def load_job_status(job_id: str) -> dict | None:
    status_file = JOBS_DIR / job_id / "status.json"
    if not status_file.exists():
        return None
    try:
        return json.loads(status_file.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def load_job_results(job_id: str) -> list[dict] | None:
    results_file = JOBS_DIR / job_id / "results.json"
    if not results_file.exists():
        return None
    try:
        return json.loads(results_file.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def load_job_type(job_id: str) -> str:
    job_file = JOBS_DIR / job_id / "job.json"
    if not job_file.exists():
        return "mpi"
    try:
        data = json.loads(job_file.read_text(encoding="utf-8"))
        return data.get("job_type", "mpi")
    except (json.JSONDecodeError, OSError):
        return "mpi"


def load_job_meta(job_id: str) -> dict:
    job_file = JOBS_DIR / job_id / "job.json"
    if not job_file.exists():
        return {}
    try:
        data = json.loads(job_file.read_text(encoding="utf-8"))
        return data.get("meta", {})
    except (json.JSONDecodeError, OSError):
        return {}


def stop_job(job_id: str):
    """実行中のジョブプロセスを停止し、ステータスを更新する"""
    import signal
    pid_file = JOBS_DIR / job_id / "pid"
    if pid_file.exists():
        try:
            pid = int(pid_file.read_text().strip())
            os.kill(pid, signal.SIGTERM)
        except (ProcessLookupError, ValueError, OSError):
            pass
    # ステータスを stopped に更新
    status = load_job_status(job_id)
    if status and status.get("state") == "running":
        status["state"] = "stopped"
        status["finished_at"] = datetime.now().isoformat()
        path = JOBS_DIR / job_id / "status.json"
        path.write_text(json.dumps(status, ensure_ascii=False, indent=2), encoding="utf-8")


def delete_job(job_id: str):
    """ジョブディレクトリを削除する"""
    import shutil
    job_dir = JOBS_DIR / job_id
    if job_dir.exists():
        shutil.rmtree(job_dir)


def list_jobs() -> list[dict]:
    """全ジョブの一覧を返す（新しい順）"""
    jobs = []
    if not JOBS_DIR.exists():
        return jobs
    for d in JOBS_DIR.iterdir():
        if not d.is_dir():
            continue
        status = load_job_status(d.name)
        meta = load_job_meta(d.name)
        if status:
            jobs.append({
                "job_id": d.name,
                "state": status.get("state", "unknown"),
                "total": status.get("total", 0),
                "completed": status.get("completed", 0),
                "errors": status.get("errors", 0),
                "started_at": status.get("started_at", ""),
                "finished_at": status.get("finished_at", ""),
                "meta": meta,
            })
    jobs.sort(key=lambda j: j.get("started_at", ""), reverse=True)
    return jobs


# ---------------------------------------------------------------------------
# 認証
# ---------------------------------------------------------------------------

def load_credentials() -> tuple[str, str]:
    """system.env からログイン情報を読み込む"""
    import os
    try:
        from env_loader import env
        login_id = env("DIGIML_LOGIN_ID", default="admin")
        login_pw = env("DIGIML_LOGIN_PW", default="admin")
    except Exception:
        login_id = os.environ.get("DIGIML_LOGIN_ID", "admin")
        login_pw = os.environ.get("DIGIML_LOGIN_PW", "admin")
    return login_id, login_pw


def check_login() -> bool:
    """ログイン画面を表示し、認証済みならTrueを返す"""
    if st.session_state.get("authenticated"):
        return True

    st.title("DigiMLab")

    valid_id, valid_pw = load_credentials()

    with st.form("login_form"):
        input_id = st.text_input("ID")
        input_pw = st.text_input("Password", type="password")
        submitted = st.form_submit_button("login", use_container_width=True)

    if submitted:
        if input_id == valid_id and input_pw == valid_pw:
            st.session_state["authenticated"] = True
            st.session_state["login_user"] = input_id
            st.rerun()
        else:
            st.error("ユーザーIDまたはパスワードが正しくありません。")

    return False


st.set_page_config(page_title="DigiMLab", page_icon="🧪", layout="wide")

# ログインチェック（未認証ならここで止まる）
if not check_login():
    st.stop()

# ---------------------------------------------------------------------------
# Streamlit ページ（認証済み）
# ---------------------------------------------------------------------------

st.title("DigiMLab - Experiment")

# サイドバー
with st.sidebar:
    st.caption(f"User: {st.session_state.get('login_user', '')}")
    if st.button("logout"):
        st.session_state["authenticated"] = False
        st.session_state.pop("login_user", None)
        st.rerun()

    # 疎通確認
    st.divider()
    models_cfg_sidebar = load_models_config()
    visible_models = [m for m in models_cfg_sidebar.get("models", []) if m.get("visible", True)]
    st.subheader("疎通確認")
    if st.button("ヘルスチェック実行"):
        for m in visible_models:
            if m.get("type") == "custom" and m.get("base_url"):
                api_base = _api_base_url(m)
                health = fetch_health.__wrapped__(api_base)  # キャッシュ無視で実行
                if health.get("status") == "ok":
                    st.success(f"{m['name']}: OK")
                else:
                    detail = health.get("detail", health.get("status", "unknown"))
                    st.error(f"{m['name']}: {detail}")

    # ジョブキュー（サイドバー）
    st.divider()
    st.subheader("ジョブキュー")
    sidebar_jobs = list_jobs()
    running_jobs = [j for j in sidebar_jobs if j["state"] == "running"]
    completed_jobs = [j for j in sidebar_jobs if j["state"] == "completed"]
    failed_jobs = [j for j in sidebar_jobs if j["state"] == "failed"]

    if running_jobs:
        for j in running_jobs:
            total = j.get("total", 1) or 1
            completed = j.get("completed", 0)
            pct = completed / total
            label = j.get("meta", {}).get("テスト名", j["job_id"])
            st.progress(pct, text=f"🔄 {label} ({completed}/{total})")
    elif not sidebar_jobs:
        st.caption("実行中のジョブはありません")

    if completed_jobs:
        st.caption(f"✅ 完了: {len(completed_jobs)}件")
    if failed_jobs:
        st.caption(f"❌ 失敗: {len(failed_jobs)}件")

tab_run, tab_results, tab_methods = st.tabs(["テスト実行", "結果閲覧・分析", "手法の理解"])

# ===== テスト実行タブ =====
with tab_run:
    settings = load_settings()
    models_cfg = load_models_config()

    col_left, col_right = st.columns([1, 1])

    with col_left:
        st.subheader("テストセット")
        suites = settings.get("test_suites", {})
        suite_options = {k: v for k, v in suites.items()}

        # テストセットソースの選択
        suite_keys = list(suite_options.keys()) + ["__upload__"]
        selected_suite = st.selectbox(
            "テストセット",
            suite_keys,
            format_func=lambda k: "Excelアップロード（任意の質問）" if k == "__upload__" else f"{k} — {suite_options[k].get('description', '')}",
        )

        # --- ジョブタイプ判定 ---
        JOB_TYPE_MAP = {
            "aipsychobench": "psycho",
            "mpi": "mpi",
            "japanese_rp_bench": "rp",
            "jp_persona": "rp",
            "culturalpersonas": "psycho",
            "rpeval": "rp",
            "your_next_token": "style",
            "characterbox": "rp",
            "personagym": "rp",
            "personallm": "mpi",
        }
        job_type = JOB_TYPE_MAP.get(selected_suite, "mpi")
        uploaded_questions = None
        upload_system_prompt = ""

        if selected_suite == "__upload__":
            job_type = "custom_qa"

            # テンプレートExcel生成
            import openpyxl as _openpyxl_qa
            from openpyxl.styles import Font as _FontQ, PatternFill as _PFillQ, Border as _BdrQ, Side as _SideQ
            _wbq = _openpyxl_qa.Workbook()
            _wsq = _wbq.active
            _wsq.title = "questions"
            _hfq = _FontQ(bold=True, color="FFFFFF")
            _hpq = _PFillQ(start_color="4472C4", end_color="4472C4", fill_type="solid")
            _tbq = _BdrQ(left=_SideQ("thin"), right=_SideQ("thin"), top=_SideQ("thin"), bottom=_SideQ("thin"))
            for col, (h, w) in enumerate([("id", 8), ("question", 60), ("category", 20)], 1):
                c = _wsq.cell(row=1, column=col, value=h)
                c.font = _hfq; c.fill = _hpq; c.border = _tbq
                _wsq.column_dimensions[c.column_letter].width = w
            for col, v in enumerate(["q001", "ここに質問文を入力", "カテゴリ（任意）"], 1):
                _wsq.cell(row=2, column=col, value=v).border = _tbq
            import io as _ioq
            _bufq = _ioq.BytesIO()
            _wbq.save(_bufq)

            st.download_button(
                "テンプレートExcelをダウンロード",
                data=_bufq.getvalue(),
                file_name="qa_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="qa_template_dl",
            )
            st.caption("必須列: `question`（質問文） / 任意列: `id`（番号）, `category`（分類）")
            uploaded_file = st.file_uploader("Excelファイルをアップロード", type=["xlsx"])
            if uploaded_file:
                try:
                    df_upload = pd.read_excel(uploaded_file)
                    if "question" not in df_upload.columns:
                        df_upload = df_upload.rename(columns={df_upload.columns[0]: "question"})
                    if "id" not in df_upload.columns:
                        df_upload["id"] = [f"q{i+1:03d}" for i in range(len(df_upload))]
                    if "category" not in df_upload.columns:
                        df_upload["category"] = ""
                    uploaded_questions = df_upload[["id", "question", "category"]].to_dict("records")
                    st.success(f"{len(uploaded_questions)}問を読み込みました")
                    st.dataframe(df_upload[["id", "question", "category"]].head(10),
                                 use_container_width=True, hide_index=True)
                except Exception as e:
                    st.error(f"ファイル読み込みエラー: {e}")

            upload_system_prompt = st.text_area(
                "システムプロンプト（任意）",
                value="",
                placeholder="回答の形式を指定したい場合などに入力してください",
            )
            suite_cfg = {"description": "Excelアップロード", "system_prompt": upload_system_prompt}
            questions = uploaded_questions or []
            total_q = len(questions)
        elif selected_suite == "your_next_token":
            job_type = "style"

            # テンプレートExcel生成
            import openpyxl as _openpyxl
            from openpyxl.styles import Font as _Font, PatternFill as _PFill, Border as _Bdr, Side as _Side, Alignment as _Align
            _wb = _openpyxl.Workbook()
            _ws = _wb.active
            _ws.title = "writing_samples"
            _hdr_font = _Font(bold=True, color="FFFFFF")
            _hdr_fill = _PFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            _thin = _Bdr(left=_Side("thin"), right=_Side("thin"), top=_Side("thin"), bottom=_Side("thin"))
            for col, (h, w) in enumerate([
                ("id", 8), ("person", 15), ("prompt", 40),
                ("sample1", 50), ("sample2", 50), ("sample3", 50), ("reference", 50),
            ], 1):
                c = _ws.cell(row=1, column=col, value=h)
                c.font = _hdr_font
                c.fill = _hdr_fill
                c.border = _thin
                _ws.column_dimensions[c.column_letter].width = w
            # サンプル行
            _example = ["s001", "著者名", "AIの未来について書いてください",
                        "私はAIの可能性に大きな期待を寄せている。（サンプル文章1）",
                        "技術と倫理のバランスが重要だ。（サンプル文章2）", "", ""]
            for col, v in enumerate(_example, 1):
                c = _ws.cell(row=2, column=col, value=v)
                c.border = _thin
            import io as _io
            _buf = _io.BytesIO()
            _wb.save(_buf)
            _template_bytes = _buf.getvalue()

            st.download_button(
                "テンプレートExcelをダウンロード",
                data=_template_bytes,
                file_name="style_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="style_template_dl",
            )
            st.caption(
                "必須列: `person`（著者名）, `prompt`（お題）, `sample1`（過去の文章サンプル1）\n\n"
                "任意列: `id`, `sample2`, `sample3`, ... , `reference`（参考文章）"
            )
            uploaded_file_style = st.file_uploader("Excelファイルをアップロード", type=["xlsx"], key="style_upload")
            if uploaded_file_style:
                try:
                    df_style = pd.read_excel(uploaded_file_style)
                    if "id" not in df_style.columns:
                        df_style["id"] = [f"s{i+1:03d}" for i in range(len(df_style))]
                    if "reference" not in df_style.columns:
                        df_style["reference"] = ""
                    # sample列を集約
                    sample_cols = [c for c in df_style.columns if c.startswith("sample")]
                    style_questions = []
                    for _, row in df_style.iterrows():
                        samples = [str(row[c]) for c in sample_cols if pd.notna(row[c]) and str(row[c]).strip()]
                        style_questions.append({
                            "id": str(row["id"]),
                            "person": str(row.get("person", "")),
                            "prompt": str(row.get("prompt", "")),
                            "writing_samples": samples,
                            "reference": str(row.get("reference", "")) if pd.notna(row.get("reference")) else "",
                        })
                    uploaded_questions = style_questions
                    st.success(f"{len(uploaded_questions)}件を読み込みました")
                    preview_df = pd.DataFrame([
                        {"ID": q["id"], "著者": q["person"], "お題": q["prompt"],
                         "サンプル数": len(q["writing_samples"])}
                        for q in uploaded_questions
                    ])
                    st.dataframe(preview_df.head(10), use_container_width=True, hide_index=True)
                except Exception as e:
                    st.error(f"ファイル読み込みエラー: {e}")

            suite_cfg = suite_options[selected_suite]
            # Excelアップロードがあればそちらを使う、なければJSONデータセット
            if uploaded_questions:
                questions = uploaded_questions
            else:
                questions = load_dataset_questions(suite_cfg.get("dataset", ""))
            total_q = len(questions)
        else:
            suite_cfg = suite_options[selected_suite]
            questions = load_dataset_questions(suite_cfg.get("dataset", ""))
            total_q = len(questions)

        if total_q > 0:
            st.caption(f"データセット: {total_q}問")

        # 件数設定（アップロード以外）
        sample_mode = "全件"
        sample_size = total_q
        random_seed = None
        if selected_suite != "__upload__" and total_q > 0:
            sample_mode = st.radio("抽出方式", ["全件", "先頭N件", "ランダムN件"], horizontal=True)
            if sample_mode != "全件":
                sample_size = st.slider("件数", min_value=1, max_value=total_q, value=min(10, total_q))
            if sample_mode == "ランダムN件":
                random_seed = st.number_input("ランダムシード (空欄=毎回異なる)", value=None, step=1)

    with col_right:
        st.subheader("接続AI (API)")
        models = [m for m in models_cfg.get("models", []) if m.get("visible", True)]
        model_names = {m["id"]: m["name"] for m in models}
        selected_model_id = st.selectbox(
            "対象モデル",
            list(model_names.keys()),
            format_func=lambda k: f"{k} — {model_names[k]}",
        )
        selected_model = next(m for m in models if m["id"] == selected_model_id)

#        st.caption(f"Type: {selected_model['type']} / URL: {selected_model.get('base_url', 'N/A')}")

        # カスタムAPIパラメータ（customタイプの場合）
        engine = None
        agent_file = None
        api_extra_params = {}
        if selected_model.get("type") == "custom":
            dp = selected_model.get("default_params", {})
            api_base = _api_base_url(selected_model)

            # --- エージェント一覧をAPIから取得 ---
            agents_list = fetch_agents(api_base)
            default_agent = dp.get("agent_file", "")
            if agents_list:
                agent_options = {a["FILE"]: f"{a['AGENT']} ({a['FILE']})" for a in agents_list}
                # デフォルトがAPI一覧にない場合は追加
                if default_agent and default_agent not in agent_options:
                    agent_options[default_agent] = f"(設定値) {default_agent}"
                agent_keys = list(agent_options.keys())
                default_idx = agent_keys.index(default_agent) if default_agent in agent_keys else 0
                agent_file = st.selectbox(
                    "エージェント",
                    agent_keys,
                    index=default_idx,
                    format_func=lambda k: agent_options[k],
                )
            else:
                agent_file = st.text_input("Agent File", value=default_agent)

            # --- エンジン一覧をAPIから取得 ---
            engines_list, engines_default = fetch_engines(api_base, agent_file or "")
            if engines_list:
                # APIデフォルト or yamlのengine設定を初期値に
                default_engine = engines_default or dp.get("engine", "")
                engine_options = ["(エージェント設定に従う)"] + engines_list
                if default_engine in engines_list:
                    default_idx = engines_list.index(default_engine) + 1
                else:
                    default_idx = 0
                selected_engine = st.selectbox("エンジン", engine_options, index=default_idx)
                engine = "" if selected_engine == "(エージェント設定に従う)" else selected_engine
            else:
                engine = st.text_input("Engine (空欄=エージェント設定に従う)", value=dp.get("engine", ""))

            with st.expander("APIパラメータ", expanded=False):
                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    api_extra_params["stream_mode"] = st.toggle("stream_mode (ストリーミング)", value=dp.get("stream_mode", True))
                    api_extra_params["memory_use"] = st.toggle("memory_use (会話履歴参照)", value=dp.get("memory_use", False))
                    api_extra_params["save_digest"] = st.toggle("save_digest (ダイジェスト保存)", value=dp.get("save_digest", False))
                    api_extra_params["magic_word_use"] = st.toggle("magic_word_use (Habit切替)", value=dp.get("magic_word_use", False))
                with col_p2:
                    api_extra_params["meta_search"] = st.toggle("meta_search (メタデータ検索)", value=dp.get("meta_search", True))
                    api_extra_params["rag_query_gene"] = st.toggle("rag_query_gene (RAGクエリ生成)", value=dp.get("rag_query_gene", True))
                    api_extra_params["web_search"] = st.toggle("web_search (Web検索)", value=dp.get("web_search", False))
                    if api_extra_params["web_search"]:
                        # Web検索エンジン一覧をAPIから取得
                        ws_engines = fetch_web_search_engines(api_base)
                        if not ws_engines:
                            ws_engines = ["OpenAI", "Perplexity", "Google"]
                        default_ws = dp.get("web_search_engine", "OpenAI")
                        ws_idx = ws_engines.index(default_ws) if default_ws in ws_engines else 0
                        api_extra_params["web_search_engine"] = st.selectbox(
                            "web_search_engine", ws_engines, index=ws_idx,
                        )
                    else:
                        api_extra_params["web_search_engine"] = dp.get("web_search_engine", "OpenAI")
                api_extra_params["situation"] = {"TIME": "", "SITUATION": ""}

        st.subheader("実行設定")
        retry_cfg = settings.get("retry", {})
        max_retries = st.number_input("最大リトライ回数", min_value=0, max_value=10,
                                       value=retry_cfg.get("max_retries", 3))
        delay_seconds = st.number_input("リクエスト間隔 (秒)", min_value=0.0, max_value=30.0,
                                         value=float(settings.get("request_delay_seconds", 0.5)), step=0.5)
        concurrency = st.number_input("並列数", min_value=1, max_value=10,
                                       value=settings.get("concurrency", 3))

    st.divider()

    if st.button("テスト実行", type="primary", use_container_width=True):
        if not questions:
            st.error(f"データセットが見つかりません: {suite_cfg.get('dataset', '')}")
        else:
            # サンプリング
            import random as _random
            if sample_mode == "ランダムN件":
                rng = _random.Random(random_seed)
                run_questions = rng.sample(questions, min(sample_size, total_q))
            elif sample_mode == "先頭N件":
                run_questions = questions[:sample_size]
            else:
                run_questions = questions

            # session_id生成（テストセット単位で固定）
            run_ts = datetime.now().strftime("%Y%m%d%H%M%S")
            login_user = st.session_state.get("login_user", "unknown")
            session_id = f"DML{run_ts}{login_user}"
            job_id = f"{selected_model_id}_{selected_suite}_{run_ts}"

            # モデル設定にUI入力値を反映
            model_for_run = dict(selected_model)
            dp = dict(model_for_run.get("default_params", {}))
            if engine is not None:
                dp["engine"] = engine
            if agent_file is not None:
                dp["agent_file"] = agent_file
            dp.update(api_extra_params)
            model_for_run["default_params"] = dp

            system_prompt = suite_cfg.get("system_prompt", "").strip()

            meta = {
                "テスト名": suite_cfg.get("description", selected_suite),
                "対象モデル": model_names.get(selected_model_id, selected_model_id),
                "engine": engine or "N/A",
                "agent_file": agent_file or "N/A",
                "実行日時": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "session_id": session_id,
                "総問題数": f"{len(run_questions)}問",
                "リクエスト間隔": f"{delay_seconds}秒",
                "最大リトライ": f"{max_retries}回",
            }

            # バックグラウンドジョブを起動
            create_job(
                job_id=job_id,
                questions=run_questions,
                model_config=model_for_run,
                system_prompt=system_prompt,
                session_id=session_id,
                delay_seconds=delay_seconds,
                max_retries=int(max_retries),
                meta=meta,
                job_type=job_type,
            )
            st.session_state["watching_job_id"] = job_id
            st.success(f"ジョブを開始しました: `{job_id}` ({len(run_questions)}問)")
            st.rerun()

    # --- ジョブキュー表示 ---
    st.divider()
    col_queue_title, col_queue_refresh = st.columns([4, 1])
    with col_queue_title:
        st.subheader("ジョブキュー")
    with col_queue_refresh:
        if st.button("🔄 最新化", key="refresh_queue"):
            st.rerun()

    jobs = list_jobs()
    if not jobs:
        st.info("実行中・完了済みのジョブはありません。")
    else:
        # 実行中ジョブがあれば自動リフレッシュ
        has_running = any(j["state"] == "running" for j in jobs)

        for j in jobs:
            job_id = j["job_id"]
            state = j["state"]
            total = j.get("total", 0)
            completed = j.get("completed", 0)
            errors = j.get("errors", 0)
            meta_info = j.get("meta", {})
            label = meta_info.get("テスト名", job_id)
            model_label = meta_info.get("対象モデル", "")
            started = j.get("started_at", "")[:19]

            finished = j.get("finished_at", "")
            duration = _format_duration(j.get("started_at", ""), finished or None)

            if state == "running":
                icon = "🔄"
                pct = completed / total if total > 0 else 0
                status_text = f"{completed}/{total} 完了 (エラー {errors}) | 経過: {duration}"
            elif state == "completed":
                icon = "✅"
                pct = 1.0
                valid = total - errors
                status_text = f"完了: 有効 {valid}/{total} (エラー {errors}) | 実行時間: {duration}"
            elif state == "stopped":
                icon = "⏹️"
                pct = completed / total if total > 0 else 0
                status_text = f"停止: {completed}/{total} まで実行済み | 実行時間: {duration}"
            elif state == "failed":
                icon = "❌"
                pct = completed / total if total > 0 else 0
                status_text = f"失敗: {j.get('error', 'unknown')}"
            else:
                icon = "⏳"
                pct = 0
                status_text = state

            with st.container(border=True):
                col_info, col_action = st.columns([4, 1])
                with col_info:
                    st.markdown(f"**{icon} {label}** — {model_label}")
                    st.caption(f"Job: {job_id} | 開始: {started}")
                    st.progress(pct, text=status_text)
                with col_action:
                    if state == "running":
                        if st.button("停止", key=f"stop_{job_id}"):
                            stop_job(job_id)
                            st.rerun()
                    if state in ("failed", "starting", "stopped", "completed"):
                        if st.button("削除", key=f"del_{job_id}"):
                            delete_job(job_id)
                            st.rerun()

        # 実行中ジョブがあれば5秒ごとに自動リフレッシュ
        if has_running:
            time.sleep(5)
            st.rerun()


# ===== 結果閲覧・分析タブ =====
with tab_results:
    st.subheader("テスト結果の閲覧・分析")
    _render_result_viewer("analysis_tab")

# ===== 手法の理解タブ =====
METHODS_DATA = [
    {
        "priority": 1,
        "name": "AIPsychoBench",
        "personality": "内面的個性（Big Five等）",
        "status": "実装済み",
        "paper_title": "AIPsychoBench: Understanding the Psychometric Differences between LLMs and Humans",
        "arxiv": "https://arxiv.org/abs/2509.16530",
        "github": "",
        "overview": "21の心理学的尺度・777問・112の下位カテゴリーで構成される多言語（8言語）の心理測定ベンチマーク。Likertスケール収集、ロールプレイプロンプト、多言語翻訳、統計分析を通じてLLMと人間の心理測定特性の違いを体系的に分析する。",
        "sections": {
            "1. Introduction": "- LLMの心理測定研究はMBTIやBig Fiveの単一尺度に偏っている\n- 英語以外の言語での検証がほぼ行われていない\n- 21尺度・8言語の包括的ベンチマークの必要性を提示",
            "2. Related Works and Motivations": "- Big Five、MBTI等の既存LLM性格評価研究のレビュー\n- 心理測定学（信頼性・妥当性）の観点からの先行研究の限界分析\n- 多尺度・多言語アプローチの動機付け",
            "3. AIPsychoBench": "- 21の心理学的尺度を6つの領域（性格・感情知能・完璧主義等）から選定\n- 777問・112下位カテゴリの質問構成\n- Likertスケール（尺度ごとに1-4〜1-10）で回答を収集\n- 8言語（英・中・日・独・仏・西・アラビア・韓）への翻訳パイプライン",
            "4. Experiment": "- GPT-4、Claude、Llama等の主要LLMに対するベンチマーク実施\n- 有効回答率の比較（モデル間で大きな差異）\n- 心理測定バイアスの検出（中心傾向バイアス、社会的望ましさバイアス）\n- 言語による回答偏差の定量分析",
            "5. Discussion and Future Works": "- LLMは人間とは異なる心理測定プロファイルを示す\n- 言語によって同じモデルでも異なる性格が表出される\n- 文化的バイアスの定量的証拠\n- 今後の拡張方向（新尺度・新言語の追加）",
            "6. Conclusion": "- 多言語心理測定ベンチマークの有用性を実証\n- LLMの「性格」は言語・プロンプトに大きく依存する",
        },
        "appendix": "",
    },
    {
        "priority": 2,
        "name": "MPI (Machine Personality Inventory)",
        "personality": "統計的個性（OCEAN）",
        "status": "実装済み",
        "paper_title": "Evaluating and Inducing Personality in Pre-trained Language Models (NeurIPS 2023 Spotlight)",
        "arxiv": "https://arxiv.org/abs/2206.07550",
        "github": "https://github.com/jianggy/MPI",
        "overview": "Big Five（OCEAN）フレームワークに基づく120問の性格診断質問紙。各質問に1〜5段階で回答させ、5因子のスコアを算出する。さらにPersonality Prompting (P2) により心理学知識とLLM知識を用いた性格誘導手法も提案。",
        "sections": {
            "1. Introduction": "- 「LLMは性格を持つか？」という問い\n- 性格の測定（評価）だけでなく誘導（操作）も研究対象\n- 人間心理学のBig Fiveモデルを機械に適用する初の体系的試み",
            "2. Related Work": "- Big Five（OCEAN）モデルの心理学的背景\n- 機械行動学（Machine Behaviour）の概念\n- NLPにおける性格分析（テキストから性格を推定する研究）との違い",
            "3. Machine Personality Inventory": "- 120問の質問紙を設計（各因子24問 × 5因子）\n- 順方向項目と逆転項目を混在させてバイアスを抑制\n- 1〜5のLikertスケールで回答、因子ごとの平均を算出\n- 既存の心理学質問紙（IPIP-NEO等）をベースに機械向けに改変",
            "4. Personality Prompting (P2)": "- 心理学的知識（各因子の行動特性記述）をプロンプトに組み込む\n- LLM内部の知識（モデルが持つ性格概念）を活用\n- 特定の性格特性を強化・抑制できることを実証\n- ビネットテスト（状況判断テスト）での検証",
            "5. Experiments": "- GPT-3、ChatGPT等の主要モデルでOCEANプロファイルを比較\n- モデルによって顕著に異なる性格プロファイルが観測される\n- 性格誘導（P2）の有効性を統計的に検証\n- 人間の性格分布との比較分析",
            "6. Conclusion": "- LLMの性格は測定可能であり、再現性がある\n- プロンプトによる性格操作は実用的に有効\n- AI安全性の観点から性格制御の重要性を指摘",
        },
        "appendix": "MPI質問項目の詳細、追加実験結果、各因子の質問分布。",
    },
    {
        "priority": 3,
        "name": "Japanese-RP-Bench",
        "personality": "表現的個性（語尾・一人称）",
        "status": "実装済み",
        "paper_title": "Japanese-RP-Bench（コミュニティプロジェクト）",
        "arxiv": "",
        "github": "https://github.com/Aratako/Japanese-RP-Bench",
        "overview": "日本語でのロールプレイにおける一貫性・自然さ・適切さ・ペルソナ精度を評価するベンチマーク。LLM-as-Judgeにより4軸で採点する。",
        "sections": {
            "概要": "- 日本語LLMのロールプレイ品質に特化したベンチマーク\n- 英語圏のRP評価では捉えられない日本語固有の要素を評価\n- 一人称（僕/私/おれ/わし等）、語尾（〜です/〜だ/〜でぇ等）、敬語レベルの一貫性",
            "評価軸": "- **consistency**: 口調・一人称・語尾がペルソナ設定と一貫しているか\n- **naturalness**: 日本語として自然で読みやすいか（文法・語彙の適切さ）\n- **relevance**: ユーザーの発話に対して的確に応じているか\n- **persona_accuracy**: ペルソナの背景・知識・価値観が応答に反映されているか",
            "データセット": "- 時代劇（江戸時代の町人等）、現代（カフェ店員等）、歴史（戦国武将風等）、ファンタジー等の多様なジャンル\n- 各ペルソナに詳細な設定（職業・口調・一人称・語尾のルール）を付与\n- 参考回答を用意し、採点の基準とする",
            "採点方式": "- GPT-4等のLLMをJudge（審査員）として使用\n- 各軸1〜5点で自動採点、採点根拠もテキストで出力\n- 複数のJudgeモデルでの採点比較も可能",
        },
        "appendix": "ペルソナ設定の一覧と参考回答例。",
    },
    {
        "priority": 4,
        "name": "JP-Persona",
        "personality": "社会的個性（役割意識）",
        "status": "実装済み",
        "paper_title": "（日本の小説キャラクターの設定再現性評価 — DigiMLab独自データセット）",
        "arxiv": "",
        "github": "",
        "overview": "日本の小説に登場するキャラクターの設定をLLMに与え、その社会的役割・価値観・行動パターンの再現性を評価する手法。",
        "sections": {
            "概要": "- 夏目漱石「坊っちゃん」、太宰治「人間失格」等の日本文学の著名キャラクターを対象\n- キャラクターの口調・価値観・行動パターンの詳細な設定をプロンプトで与える\n- 状況に応じた応答がキャラクターらしいかを評価",
            "評価観点": "- **役割意識の一貫性**: キャラクターの社会的立場（教師・学生・武将等）に応じた言動か\n- **社会的文脈の理解**: 時代背景・人間関係を踏まえた応答か\n- **口調の再現性**: 一人称・語尾・方言等が設定通りか\n- **価値観の反映**: キャラクターの信念・思想が応答に表れているか",
            "DigiMLabでの実装": "- 10作品のキャラクター（坊っちゃん、ジョバンニ、葉蔵、龍馬等）を収録\n- Japanese-RP-Benchと同じ4軸Judge採点が利用可能\n- 追加のキャラクターはデータセットを編集して拡張可能",
        },
        "appendix": "",
    },
    {
        "priority": 5,
        "name": "CulturalPersonas",
        "personality": "文化的個性（空気を読む等）",
        "status": "実装済み",
        "paper_title": "Can LLMs Express Personality Across Cultures? Introducing CulturalPersonas for Evaluating Trait Alignment",
        "arxiv": "https://arxiv.org/abs/2506.05670",
        "github": "",
        "overview": "6カ国（米・英・独・日・中・印）の文化的背景に基づく500問のシナリオベースMCQで、Big Five特性の文化間差異を評価。",
        "sections": {
            "1. Introduction": "- LLMの性格表出は英語・西洋文化を前提とした評価が主流\n- 同じBig Five特性でも文化によって「望ましい行動」が異なる\n- 例: 日本文化圏の「協調性」と米国文化圏の「協調性」は質が異なる",
            "2. Background": "- Big Fiveモデルの文化横断的な研究の歴史\n- Hofstedeの文化次元理論との関連\n- LLMの訓練データに含まれる文化的バイアスの問題",
            "3. CulturalPersonas": "- 6カ国（米・英・独・日・中・印）の文化的に根拠のあるシナリオを作成\n- 各国×Big Five 5因子の組み合わせで500問のMCQを設計\n- シナリオは各国の文化的専門家の監修を受けて作成\n- 例: 日本のシナリオでは「空気を読む」「和を重んじる」等の要素",
            "4. Experimental Setup": "- GPT-4、Claude、Llama等の主要LLMで実験\n- 各モデルに6カ国のペルソナを割り当てて回答させる\n- 実際の各国の心理学的データとの比較",
            "5. Evaluating Multi-Cultural Personalities in LLMs": "- LLMは西洋文化圏の性格パターンを優先的に表出する傾向\n- 日本文化圏では協調性・神経症傾向が高く出る傾向\n- モデルによって文化的感度に大きな差異\n- 訓練データの言語分布が文化的バイアスに直結",
            "6. Conclusion & Future Work": "- 文化的に公平なAI評価の必要性を実証\n- 文化的ステレオタイプを再現するリスクへの警鐘\n- 多文化対応AI開発のための評価ガイドラインの提案",
        },
        "appendix": "A. データセット生成 / B. 人間による検証 / C. 性格評価手法 / D. 追加結果 / E. 追加アブレーション。",
    },
    {
        "priority": 6,
        "name": "RPEval",
        "personality": "倫理的個性",
        "status": "実装済み",
        "paper_title": "Role-Playing Evaluation for Large Language Models",
        "arxiv": "https://arxiv.org/abs/2505.13157",
        "github": "https://github.com/yelboudouri/RPEval",
        "overview": "感情理解・意思決定・道徳的一貫性・キャラクター知識の4軸で、LLMのロールプレイ能力を包括的に評価するフレームワーク。",
        "sections": {
            "1. Introduction": "- 既存のRP評価は「キャラクターらしさ」の1軸に偏っている\n- 感情・倫理・知識・意思決定を含む多軸評価の必要性\n- 9,018件の大規模評価データセットを構築",
            "2. Design Considerations": "- **感情理解（emotion）**: キャラクターの感情状態を正しく表現できるか\n- **意思決定（decision）**: キャラクターの価値観に基づく判断ができるか\n- **道徳整合性（moral）**: キャラクターの倫理観と矛盾しない行動か\n- **キャラクター一貫性（in-character）**: 設定から逸脱しない応答か",
            "3. Benchmark Construction": "- 映画・小説・ゲーム等から多様なキャラクターを収集\n- 各キャラクターにsystem promptとしてペルソナ設定を付与\n- 3タイプ（emotion/decision/in-character）の評価シナリオを構築\n- 人間アノテーターによる品質検証",
            "4. Evaluation Results": "- 主要LLMの4軸比較: 感情理解はGPT-4が強い等の傾向\n- 意思決定と道徳整合性は全モデルで課題が残る\n- キャラクター一貫性はモデルサイズに比例する傾向",
            "5. Conclusion": "- 多軸評価により各モデルの強み・弱みが明確に\n- ロールプレイAI開発における具体的な改善指針を提示",
        },
        "appendix": "A. シナリオ例（感情理解、意思決定、道徳整合性、キャラクター一貫性）。",
    },
    {
        "priority": 7,
        "name": "Your Next Token Prediction (YNTP-100)",
        "personality": "言語的個性",
        "status": "実装済み",
        "paper_title": "YNTP-100: A Benchmark for Your Next Token Prediction with 100 People",
        "arxiv": "https://arxiv.org/abs/2510.14398",
        "github": "",
        "overview": "100人の多言語マルチターン対話データから、個人の執筆スタイル（Substance=何を言うか、Style=どう言うか）の再現度を評価するベンチマーク。",
        "sections": {
            "1. Introduction": "- 「この人ならこう書く」をLLMに再現させるタスクの定義\n- 既存のパーソナライゼーション研究は嗜好推薦に偏っている\n- 言語スタイル自体の再現に焦点を当てた初のベンチマーク",
            "2. Related Work": "- 著者識別（Authorship Attribution）: テキストから著者を特定する研究\n- スタイル転写（Style Transfer）: 文体を変換する研究\n- パーソナライズドLLM: ユーザーの好みに合わせたモデル調整",
            "3. YNTP-100 Benchmark": "- 100人の参加者が5日間にわたりMBTI設定のNPCと多言語（英日中）で対話\n- **Substance（内容）**: 何を話すか — 話題選択・主張の方向性\n- **Style（文体）**: どう話すか — 語彙・語尾・文長・口調\n- 各人物の過去4日分の対話を履歴として与え、5日目の応答を予測",
            "4. Experiments Setup": "- **外部アライメント**: プロンプトに履歴を与えてスタイルを模倣させる（ICL）\n- **内部アライメント**: LoRA等でファインチューニング\n- 複数のモデルサイズ（7B〜70B）で比較",
            "5. Results and Analysis": "- ファインチューニングはSubstance・Style両方で優位\n- プロンプトベース（ICL）はStyleの再現に一定の効果\n- モデルサイズが大きいほどスタイル再現精度が向上\n- 日本語は英語・中国語より再現が難しい傾向",
            "6. Conclusion": "- 個人スタイル再現はLLMの重要な能力として今後注目される\n- ファインチューニングが最も有効だが、プロンプト設計でも改善可能",
        },
        "appendix": "A. 既存データセット一覧 / B. データ収集プロトコル / C. 評価指標の数式 / D. 元実験結果 / E. プロンプトとハイパーパラメータ / F. ケーススタディ。",
    },
    {
        "priority": 8,
        "name": "CharacterBox",
        "personality": "適応的個性",
        "status": "実装済み",
        "paper_title": "CharacterBox: Evaluating the Role-Playing Capabilities of LLMs in Text-Based Virtual Worlds (NAACL 2025)",
        "arxiv": "https://arxiv.org/abs/2412.05631",
        "github": "",
        "overview": "テキストベースの仮想世界における状況変化下で、LLMがキャラクターの性格をどの程度維持・適応できるかを評価。キャラクターエージェントとナレーターエージェントによるシミュレーション設計で行動軌跡の細粒度分析を行う。",
        "sections": {
            "1. Introduction": "- 従来のRP評価は静的なQ&A形式で、動的な状況変化への対応を測れない\n- 「ストレス下でもキャラクターを維持できるか？」という問い\n- テキスト仮想世界でのシミュレーション評価を提案",
            "2. Related Work": "- インタラクティブフィクション・テキストアドベンチャーの歴史\n- LLMベースのキャラクター生成（Character.ai等）の研究\n- 動的環境でのエージェント評価手法",
            "3. CharacterBox Framework": "- **キャラクターエージェント**: LLMがキャラクターとして行動する\n- **ナレーターエージェント**: 状況変化イベント（ストレス・道徳的ジレンマ等）を生成する\n- Harry Potter等の10作品からキャラクター・シーンを抽出\n- 行動軌跡（行動の連続記録）を細粒度で分析",
            "4. Experiments": "- 主要LLMの性格維持度・適応度・合理性を比較\n- 軽量モデル（CharacterNR: ナレーター、CharacterRM: 報酬モデル）の有効性検証\n- ストレス状況での性格崩れパターンの分析\n- 道徳的ジレンマでのキャラクター一貫性の検証",
            "5. Conclusion": "- 大規模モデルほど状況変化に強い傾向\n- 動的評価は静的評価では発見できない弱点を明らかにする\n- ナレーターモデルによる自動シナリオ生成の有効性",
        },
        "appendix": "詳細なシーン設定・評価基準、行動軌跡の例。",
    },
    {
        "priority": 9,
        "name": "PersonaGym",
        "personality": "知的個性",
        "status": "実装済み",
        "paper_title": "PersonaGym: Evaluating Persona Agents and LLMs (EMNLP Findings 2025)",
        "arxiv": "https://arxiv.org/abs/2407.18416",
        "github": "https://github.com/vsamuel2003/PersonaGym",
        "overview": "200のペルソナ×150の動的環境で10,000問を自動生成し、ペルソナエージェントの意思決定プロセスを包括的に評価するフレームワーク。",
        "sections": {
            "1. Introduction": "- ペルソナエージェント（特定の人物になりきるAI）の品質をどう測るか\n- 既存評価は限定的な質問セットに依存し、網羅性が不足\n- 動的な環境・質問生成による包括的評価を提案",
            "2. Evaluation Tasks": "- **Expected Action**: ペルソナに期待される行動の予測\n- **Toxicity**: 有害な応答の生成リスク\n- **Linguistic Habits**: 語彙・口調の一貫性\n- **Persona Consistency**: 設定からの逸脱検出\n- **Action Justification**: 行動の根拠が設定と整合するか",
            "3. PersonaGym": "- 200のペルソナ定義（職業・性格・背景等）を用意\n- 150の動的環境（日常・緊急・倫理的ジレンマ等）を設計\n- 環境選択フェーズ: ペルソナに適した環境を動的に選択\n- 質問生成フェーズ: 環境に応じた質問を自動生成\n- 合計10,000問を自動生成",
            "4. Experiments": "- GPT-4、Claude、Llama等で比較\n- GPT-4が全体的に高スコアだが、Toxicityでは差が小さい\n- 環境の複雑さが増すとペルソナ崩れが顕著に\n- 小型モデルはLinguistic Habitsで特に弱い",
            "5. Human Evaluation": "- 人手評価者200名による妥当性検証\n- 自動評価と人手評価の相関が高いことを確認\n- LLM-as-Judgeの信頼性を統計的に実証",
            "6. Related Work": "- ペルソナ対話システム（Persona-Chat等）の歴史\n- エージェント評価ベンチマーク（AgentBench等）との比較",
            "7. Conclusion": "- 動的環境での評価が静的評価より多くの弱点を発見\n- ペルソナエージェント開発における5つの改善ポイントを提示",
        },
        "appendix": "A. プロンプト / B. 環境設定 / C. 定性例 / D. ペルソナ一覧 / E. 定式化の記法 / F. 有意差検定。",
    },
    {
        "priority": 10,
        "name": "PersonaLLM",
        "personality": "知覚的個性",
        "status": "実装済み",
        "paper_title": "PersonaLLM: Investigating the Ability of Large Language Models to Express Personality Traits (NAACL Findings 2024)",
        "arxiv": "https://arxiv.org/abs/2305.02547",
        "github": "https://github.com/hjian42/PersonaLLM",
        "overview": "Big Fiveに基づくペルソナをLLMに割り当て、44項目のBFI性格検査と創作文章（エッセイ）により、人手評価と自動評価の両面からLLMの性格表出能力を検証。",
        "sections": {
            "1. Introduction": "- 「LLMに性格を与えたら、人間はそれを読み取れるか？」という問い\n- Big Fiveの5因子×高低の10パターンのペルソナを設定\n- BFI質問紙への回答とエッセイ執筆の2つのタスクで検証",
            "2. Experiment Design": "- GPT-3.5/GPT-4に10パターンのペルソナプロンプトを与える\n- **BFI-44テスト**: 44項目の性格検査に回答させ、自己認識を測定\n- **物語生成**: 各ペルソナでショートストーリーを執筆させる\n- 各パターン10回反復して再現性を確認",
            "3. Results": "- **RQ1（BFI回答）**: LLMは指示された因子を概ね高く/低く回答できる\n- **RQ2（言語パターン）**: 外向性が高いペルソナは語彙が豊富、神経症傾向が高いと否定語が増加\n- **RQ3（物語評価）**: 人間評価者はAIの性格を一定程度正しく識別\n- **RQ4（性格知覚）**: 神経症傾向の表現が最も難しく、協調性は比較的容易",
            "4. Related Work": "- 性格心理学における言語使用の研究（Pennebaker等）\n- LLMエージェントの行動分析\n- NLPにおけるテキストからの性格推定",
            "5. Conclusion": "- LLMは概ね指示された性格を文章に表出できる\n- 神経症傾向（不安・抑うつ等）の表現が最も困難\n- 人間評価者はAI生成テキストから性格を読み取れるが、完璧ではない",
        },
        "appendix": "A. LLM生成物語例 / B. 物語コメント / C. 性格評定 / D. 物語評価詳細 / E. LLM評価者 / F. LLaMA 2のBFIスコア / G. 性格特性の追加結果。",
    },
]

with tab_methods:
    st.subheader("評価手法の解説")
    st.caption("DigiMLabで使用・参照している10種類のパーソナリティ評価手法を、元論文に基づいて解説します。")

    for m in METHODS_DATA:
        with st.expander(f"**{m['priority']}. {m['name']}** — {m['personality']}", expanded=False):
            # 基本情報
            st.markdown(f"### {m['name']}")
            st.markdown(f"**論文:** {m['paper_title']}")
            col_link1, col_link2, col_link3 = st.columns(3)
            if m["arxiv"]:
                col_link1.markdown(f"[arXiv]({m['arxiv']})")
            if m["github"]:
                col_link2.markdown(f"[GitHub]({m['github']})")
            # PDF ダウンロード
            pdf_map = {
                "AIPsychoBench": "AIPsychoBench.pdf",
                "MPI (Machine Personality Inventory)": "MPI.pdf",
                "CulturalPersonas": "CulturalPersonas.pdf",
                "RPEval": "RPEval.pdf",
                "Your Next Token Prediction (YNTP-100)": "YNTP-100.pdf",
                "CharacterBox": "CharacterBox.pdf",
                "PersonaGym": "PersonaGym.pdf",
                "PersonaLLM": "PersonaLLM.pdf",
            }
            pdf_file = pdf_map.get(m["name"], "")
            pdf_path = Path(__file__).parent / "datasets" / "papers" / pdf_file if pdf_file else None
            if pdf_path and pdf_path.is_file():
                with open(pdf_path, "rb") as _fpdf:
                    col_link3.download_button(
                        "論文PDF",
                        data=_fpdf.read(),
                        file_name=pdf_path.name,
                        mime="application/pdf",
                        key=f"paper_pdf_{m['name']}",
                    )
            st.markdown(f"**評価される個性:** {m['personality']}")
            st.divider()

            # 概要
            st.markdown("#### 概要")
            st.write(m["overview"])

            # セクション解説
            if m.get("sections"):
                st.markdown("#### 論文の構成")
                for sec_title, sec_desc in m["sections"].items():
                    st.markdown(f"**{sec_title}**")
                    st.markdown(sec_desc)

            # Appendix
            if m.get("appendix"):
                st.markdown("#### Appendix")
                st.caption(m["appendix"])
